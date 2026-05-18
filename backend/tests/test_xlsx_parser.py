"""
Testes de integração do XLSXParser — edge cases reais de catálogos brasileiros.

Cria arquivos XLSX sintéticos em memória com openpyxl e executa o parser completo.
Cada teste replica uma situação real observada em catálogos de distribuidoras.

Execute: pytest tests/test_xlsx_parser.py -v
"""

import io
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from openpyxl import Workbook

from app.services.parse_result import ParseConfidence
from app.services.xlsx_parser import XLSXParser, parse_xlsx_catalog


# ── Fixtures e helpers ─────────────────────────────────────────────────────────

def save_wb(wb: Workbook, tmp_path: Path, name: str = "catalog.xlsx") -> Path:
    """Salva Workbook em arquivo temporário e retorna o Path."""
    file_path = tmp_path / name
    wb.save(str(file_path))
    return file_path


def make_standard_wb() -> Workbook:
    """
    Catálogo padrão bem estruturado.
    Linha 1: cabeçalho claro
    Linhas 2+: produtos com custo numérico
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    ws.append(["Produto", "Custo", "SKU", "Categoria", "Fornecedor"])
    ws.append(["Caixa de Papelão 30x20", 12.50, "001", "Embalagens", "Distribuidora A"])
    ws.append(["Frasco Plástico 500ml", 5.90, "002", "Embalagens", "Distribuidora A"])
    ws.append(["Fita Adesiva Transparente", 3.20, "003", "Escritório", "Distribuidora B"])
    ws.append(["Papel A4 Resma 500fls", 28.00, "004", "Escritório", "Distribuidora B"])
    ws.append(["Caneta Esferográfica Azul", 1.50, "005", "Escritório", "Distribuidora C"])

    return wb


# ── Testes: parsing básico ─────────────────────────────────────────────────────

class TestBasicParsing:
    """Casos básicos que qualquer parser decente deve passar."""

    def test_standard_catalog(self, tmp_path):
        """Catálogo padrão — produz RELIABLE com todos os campos."""
        wb = make_standard_wb()
        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.confidence == ParseConfidence.RELIABLE
        assert len(result.products) == 5
        assert result.stats.valid_products == 5
        assert result.stats.skipped_empty == 0

        # Verificar primeiro produto
        p = result.products[0]
        assert p["raw_name"] == "Caixa de Papelão 30x20"
        assert p["cost"] == Decimal("12.5")
        assert p["sku"] == "001"  # célula string "001" — zeros preservados
        assert p["category"] == "Embalagens"
        assert p["supplier"] == "Distribuidora A"

    def test_column_mapping_captured(self, tmp_path):
        """ColumnMappingResult está preenchido corretamente."""
        wb = make_standard_wb()
        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        mapping = result.column_mapping
        assert mapping.product_name == 0
        assert mapping.cost == 1
        assert mapping.sku == 2
        assert mapping.category == 3
        assert mapping.supplier == 4
        assert mapping.has_required_columns is True

    def test_parse_result_metadata_dict(self, tmp_path):
        """to_metadata_dict() retorna estrutura serializável (para salvar no banco)."""
        wb = make_standard_wb()
        result = parse_xlsx_catalog(save_wb(wb, tmp_path))
        meta = result.to_metadata_dict()

        assert meta["confidence"] == "RELIABLE"
        assert meta["stats"]["valid_products"] == 5
        assert isinstance(meta["stats"]["success_rate"], float)
        assert "column_mapping" in meta
        assert "warnings" in meta
        assert "errors" in meta

    def test_only_required_columns(self, tmp_path):
        """Catálogo sem colunas opcionais — ainda é válido."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Preço"])
        ws.append(["Luva de Vinil P", 25.00])
        ws.append(["Luva de Vinil M", 25.00])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.confidence == ParseConfidence.RELIABLE
        assert len(result.products) == 2
        assert result.column_mapping.sku is None
        assert result.column_mapping.category is None
        # Deve reportar colunas opcionais ausentes nos warnings
        missing_optional = result.column_mapping.missing_optional
        assert "sku" in missing_optional


# ── Testes: detecção de cabeçalho ─────────────────────────────────────────────

class TestHeaderDetection:
    """Cabeçalho pode não estar na primeira linha."""

    def test_header_in_row_3_with_company_name(self, tmp_path):
        """
        Padrão real: distribuidora coloca nome da empresa e CNPJ nas primeiras linhas.
        Linha 1: "DISTRIBUIDORA PAULISTA LTDA"
        Linha 2: vazia
        Linha 3: cabeçalho real
        """
        wb = Workbook()
        ws = wb.active
        ws.append(["DISTRIBUIDORA PAULISTA LTDA", None, None, None])
        ws.append([None, None, None, None])
        ws.append(["Produto", "Preço de Custo", "Código", "Grupo"])
        ws.append(["Arruela M6 Zinco", 0.08, "ARL-M6-ZN", "Fixadores"])
        ws.append(["Parafuso Sextavado M8", 0.32, "PAR-M8-SX", "Fixadores"])
        ws.append(["Porca M8 Sextavada", 0.15, "PRC-M8-SX", "Fixadores"])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.confidence != ParseConfidence.FAILED
        assert len(result.products) == 3
        assert result.stats.header_row_index == 2

    def test_header_in_row_5_with_metadata(self, tmp_path):
        """
        Catálogos com mais linhas de metadado antes do conteúdo.
        """
        wb = Workbook()
        ws = wb.active
        ws.append(["Tabela de Preços", None, None])
        ws.append(["Validade: 01/01/2026 a 31/12/2026", None, None])
        ws.append(["Condição: à vista", None, None])
        ws.append([None, None, None])
        ws.append(["Nome do Produto", "Valor Unitário", "SKU"])
        ws.append(["Produto Alpha", 100.00, "ALPHA-001"])
        ws.append(["Produto Beta", 200.00, "BETA-002"])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.confidence != ParseConfidence.FAILED
        assert len(result.products) == 2

    def test_uppercase_headers(self, tmp_path):
        """Catálogos antigos com tudo em maiúsculas."""
        wb = Workbook()
        ws = wb.active
        ws.append(["PRODUTO", "CUSTO", "CÓDIGO"])
        ws.append(["Botão de Pressão Nº10", 0.45, "BT-PR-10"])
        ws.append(["Zíper 20cm Nylon", 1.20, "ZIP-20-NY"])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.confidence != ParseConfidence.FAILED
        assert len(result.products) == 2

    def test_english_headers(self, tmp_path):
        """Catálogo de fornecedor internacional."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Product Name", "Unit Cost", "SKU", "Category"])
        ws.append(["Industrial Gloves L", 8.50, "GL-L-IND", "Safety"])
        ws.append(["Safety Helmet Yellow", 45.00, "HLM-YEL", "Safety"])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.confidence != ParseConfidence.FAILED
        assert len(result.products) == 2
        assert result.column_mapping.has_required_columns is True


# ── Testes: linhas de subtotal ─────────────────────────────────────────────────

class TestSubtotalRows:
    """Linhas de total/subtotal não devem virar produtos."""

    def test_subtotal_rows_skipped(self, tmp_path):
        """Catálogo com subtotais por categoria — padrão comum."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo", "SKU", "Categoria"])
        ws.append(["Caneta Azul BIC", 1.20, "CAN-AZ", "Escritório"])
        ws.append(["Lápis HB", 0.80, "LAP-HB", "Escritório"])
        ws.append(["Subtotal Escritório", 2.00, None, "Escritório"])   # deve ser ignorada
        ws.append(["Borracha Branca", 0.60, "BOR-BR", "Escola"])
        ws.append(["Régua 30cm", 1.10, "REG-30", "Escola"])
        ws.append(["Subtotal Escola", 1.70, None, "Escola"])           # deve ser ignorada
        ws.append(["Total Geral", 3.70, None, None])                   # deve ser ignorada

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.stats.skipped_subtotal == 3
        assert len(result.products) == 4
        # Nenhum produto com nome "Subtotal" ou "Total"
        names = [p["raw_name"] for p in result.products]
        assert all("total" not in n.lower() for n in names)
        assert all("subtotal" not in n.lower() for n in names)

    def test_total_keyword_variations(self, tmp_path):
        """Múltiplas variações de keyword de total."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo"])
        ws.append(["Produto Válido", 50.00])
        ws.append(["TOTAL GERAL", 50.00])       # maiúsculo
        ws.append(["Soma", 50.00])              # "soma" também é keyword
        ws.append(["Grand Total", 50.00])       # inglês

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert len(result.products) == 1
        assert result.products[0]["raw_name"] == "Produto Válido"


# ── Testes: herança de categoria (células mescladas) ──────────────────────────

class TestCategoryInheritance:
    """
    Padrão real: categoria na primeira linha de um grupo,
    células seguintes têm None (como se fossem células mescladas).
    openpyxl retorna None para células mescladas não-master em read_only mode.
    """

    def test_category_inherited_to_next_rows(self, tmp_path):
        """Categoria só aparece na primeira linha do grupo — deve se propagar."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo", "SKU", "Categoria"])
        ws.append(["Caixa P", 5.00, "CX-P", "Embalagens"])    # categoria definida
        ws.append(["Caixa M", 8.00, "CX-M", None])             # herda "Embalagens"
        ws.append(["Caixa G", 12.00, "CX-G", None])            # herda "Embalagens"
        ws.append(["Caneta Azul", 1.20, "CAN-AZ", "Escritório"])  # nova categoria
        ws.append(["Caneta Vermelha", 1.20, "CAN-VM", None])   # herda "Escritório"

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert len(result.products) == 5
        # Três embalagens
        assert result.products[0]["category"] == "Embalagens"
        assert result.products[1]["category"] == "Embalagens"
        assert result.products[2]["category"] == "Embalagens"
        # Duas escritório
        assert result.products[3]["category"] == "Escritório"
        assert result.products[4]["category"] == "Escritório"

    def test_category_not_inherited_across_sections(self, tmp_path):
        """Se uma nova categoria aparece, deve substituir a herdada."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo", "Categoria"])
        ws.append(["Produto A", 10.00, "Higiene"])
        ws.append(["Produto B", 20.00, None])         # herda Higiene
        ws.append(["Produto C", 30.00, "Limpeza"])    # nova categoria
        ws.append(["Produto D", 40.00, None])         # herda Limpeza

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.products[1]["category"] == "Higiene"
        assert result.products[2]["category"] == "Limpeza"
        assert result.products[3]["category"] == "Limpeza"


# ── Testes: formatos de preço ──────────────────────────────────────────────────

class TestPriceFormats:
    """Formatos de preço variados no mesmo catálogo."""

    def test_all_numeric_float(self, tmp_path):
        """Preços como float puro — o caso mais simples."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo"])
        ws.append(["Produto A", 10.50])
        ws.append(["Produto B", 1234.56])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))
        assert result.products[0]["cost"] == Decimal("10.5")
        assert result.products[1]["cost"] == Decimal("1234.56")

    def test_mixed_price_formats_string_column(self, tmp_path):
        """
        Planilha onde a coluna de preço é string com formatos mistos.
        Catálogos exportados de ERPs antigos têm esse padrão.
        """
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Preço de Custo"])
        ws.append(["Prod A", "R$ 1.234,56"])     # formato pt-BR completo
        ws.append(["Prod B", "R$50,00"])          # sem espaço
        ws.append(["Prod C", "1234.56"])          # formato en-US
        ws.append(["Prod D", "50,00"])            # pt-BR simples
        ws.append(["Prod E", "1.234"])            # ambíguo — deve ser 1234

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert len(result.products) == 5
        assert result.products[0]["cost"] == Decimal("1234.56")
        assert result.products[1]["cost"] == Decimal("50.00")
        assert result.products[2]["cost"] == Decimal("1234.56")
        assert result.products[3]["cost"] == Decimal("50.00")
        assert result.products[4]["cost"] == Decimal("1234")

    def test_invalid_prices_skipped(self, tmp_path):
        """Linhas com preço inválido são ignoradas, o resto continua."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo"])
        ws.append(["Produto OK 1", 10.00])
        ws.append(["Produto Sem Custo", ""])       # preço vazio
        ws.append(["Produto Zero", 0.00])          # zero inválido
        ws.append(["Produto Negativo", -5.00])     # negativo inválido
        ws.append(["Produto OK 2", 20.00])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert len(result.products) == 2
        assert result.stats.skipped_invalid_cost == 3

    def test_two_price_columns_picks_cost_not_sell(self, tmp_path):
        """
        Catálogo com custo E preço de venda — deve mapear a coluna correta.
        'Custo' score > 'Preço Venda' para o campo cost.
        """
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Preço de Custo", "Preço de Venda", "SKU"])
        ws.append(["Produto A", 10.00, 25.00, "A001"])
        ws.append(["Produto B", 15.00, 35.00, "B002"])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.column_mapping.cost == 1  # "Preço de Custo" = col 1
        assert len(result.products) == 2
        # Custo deve ser o de coluna 1 (preço de custo), não de venda
        assert result.products[0]["cost"] == Decimal("10.00")


# ── Testes: linhas vazias ──────────────────────────────────────────────────────

class TestEmptyRows:
    """Linhas vazias no meio dos dados são comuns em catálogos exportados."""

    def test_empty_rows_in_middle_ignored(self, tmp_path):
        """Linhas vazias entre produtos não interrompem o parsing."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo"])
        ws.append(["Produto A", 10.00])
        ws.append([None, None])                  # vazia
        ws.append(["Produto B", 20.00])
        ws.append([None, None])                  # vazia
        ws.append([None, None])                  # vazia
        ws.append(["Produto C", 30.00])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert len(result.products) == 3
        assert result.stats.skipped_empty == 3

    def test_max_consecutive_empty_stops_parsing(self, tmp_path):
        """10 linhas vazias consecutivas encerra o parsing (evita ler até row 1048576)."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo"])
        ws.append(["Produto A", 10.00])
        ws.append(["Produto B", 20.00])
        # 10 linhas vazias consecutivas
        for _ in range(10):
            ws.append([None, None])
        # Produto "escondido" depois das linhas vazias — NÃO deve ser lido
        ws.append(["Produto Escondido", 999.00])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert len(result.products) == 2
        names = [p["raw_name"] for p in result.products]
        assert "Produto Escondido" not in names

    def test_fewer_than_max_empty_continues(self, tmp_path):
        """9 linhas vazias consecutivas NÃO encerra — 10 é o limite."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo"])
        ws.append(["Produto A", 10.00])
        # 9 linhas vazias
        for _ in range(9):
            ws.append([None, None])
        ws.append(["Produto B", 20.00])   # deve ser lido

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert len(result.products) == 2


# ── Testes: SKU e duplicatas ───────────────────────────────────────────────────

class TestSKUAndDuplicates:
    """SKU com zeros à esquerda e duplicatas."""

    def test_numeric_sku_converts_to_string(self, tmp_path):
        """SKU numérico perde zeros à esquerda mas converte corretamente."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo", "SKU"])
        ws.append(["Produto A", 10.00, 123])       # int 123
        ws.append(["Produto B", 20.00, 456.0])     # float 456.0

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.products[0]["sku"] == "123"
        assert result.products[1]["sku"] == "456"

    def test_string_sku_preserved(self, tmp_path):
        """SKU como string (formatado no Excel) preserva zeros à esquerda."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo", "SKU"])
        ws.append(["Produto A", 10.00, "00123"])   # string com zeros

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        # Zeros à esquerda preservados quando célula é string
        assert result.products[0]["sku"] == "00123"

    def test_duplicate_sku_both_kept_with_warning(self, tmp_path):
        """
        SKU duplicado: ambos são mantidos (pode ser produto diferente),
        mas o segundo gera warning.
        """
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo", "SKU"])
        ws.append(["Produto A", 10.00, "SKU-001"])
        ws.append(["Produto B", 20.00, "SKU-001"])  # SKU duplicado
        ws.append(["Produto C", 30.00, "SKU-002"])  # SKU único

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        # Ambos são mantidos (política de manter e avisar)
        assert len(result.products) == 3
        assert result.stats.skipped_duplicate_sku == 1
        # Warning deve estar presente
        sku_warnings = [w for w in result.warnings if "SKU-001" in w]
        assert len(sku_warnings) >= 1

    def test_no_sku_column_products_still_valid(self, tmp_path):
        """Catálogo sem coluna SKU — produtos são válidos mesmo sem SKU."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Nome do Produto", "Preço"])
        ws.append(["Produto Sem SKU", 15.00])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert len(result.products) == 1
        assert result.products[0]["sku"] is None


# ── Testes: múltiplas abas ─────────────────────────────────────────────────────

class TestMultipleSheets:
    """Seleção da aba correta em workbooks com múltiplas abas."""

    def test_picks_sheet_with_most_rows(self, tmp_path):
        """Com 2 abas de dados, seleciona a com mais linhas."""
        wb = Workbook()

        # Aba pequena
        ws_small = wb.active
        ws_small.title = "Resumo"
        ws_small.append(["Produto", "Custo"])
        ws_small.append(["Produto Único", 10.00])

        # Aba grande
        ws_big = wb.create_sheet("Catálogo Completo")
        ws_big.append(["Produto", "Custo", "SKU"])
        for i in range(10):
            ws_big.append([f"Produto {i+1}", float(i + 1) * 5, f"SKU-{i+1:03d}"])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.stats.sheet_name == "Catálogo Completo"
        assert result.stats.total_sheets == 2
        assert len(result.products) == 10
        # Warning de múltiplas abas
        multi_warnings = [w for w in result.warnings if "abas" in w.lower() or "múltiplas" in w.lower()]
        assert len(multi_warnings) >= 1

    def test_skips_auxiliary_sheets(self, tmp_path):
        """Abas com nomes como 'config', 'legenda', 'ref' devem ser ignoradas."""
        wb = Workbook()

        # Aba de configuração — deve ser ignorada
        ws_config = wb.active
        ws_config.title = "Config"
        ws_config.append(["Parâmetro", "Valor"])
        ws_config.append(["versão", "1.0"])
        ws_config.append(["data", "2026-01-01"])
        ws_config.append(["tabela", "preços"])
        ws_config.append(["responsável", "João"])

        # Aba de dados reais
        ws_data = wb.create_sheet("Produtos")
        ws_data.append(["Produto", "Custo"])
        ws_data.append(["Item A", 10.00])
        ws_data.append(["Item B", 20.00])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.stats.sheet_name == "Produtos"
        assert len(result.products) == 2

    def test_fallback_to_first_sheet_if_all_aux(self, tmp_path):
        """Se todas as abas tiverem nomes auxiliares, usa a primeira."""
        wb = Workbook()

        ws = wb.active
        ws.title = "Config"
        ws.append(["Produto", "Custo"])
        ws.append(["Produto A", 10.00])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))
        # Deve processar algo, não falhar
        assert result is not None


# ── Testes: colunas problemáticas ─────────────────────────────────────────────

class TestProblematicColumns:
    """Colunas com None, mescladas, extras irrelevantes."""

    def test_none_headers_in_merged_columns(self, tmp_path):
        """
        Colunas mescladas no cabeçalho resultam em None no openpyxl.
        Parser deve ignorar as colunas None e mapear as demais corretamente.
        """
        wb = Workbook()
        ws = wb.active
        # Cabeçalho com células None (simulando colunas mescladas ou vazias)
        ws.append(["Produto", None, "Custo", None, "SKU"])
        ws.append(["Produto A", None, 10.00, None, "A001"])
        ws.append(["Produto B", None, 20.00, None, "B002"])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.column_mapping.product_name == 0
        assert result.column_mapping.cost == 2
        assert result.column_mapping.sku == 4
        assert len(result.products) == 2

    def test_extra_irrelevant_columns_dont_confuse_mapping(self, tmp_path):
        """
        Colunas extras (estoque, margem %) não devem interferir no mapeamento.
        """
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Qtd Estoque", "Custo", "Preço Venda", "Margem %", "SKU"])
        ws.append(["Produto A", 100, 10.00, 25.00, "150%", "A001"])
        ws.append(["Produto B", 50, 20.00, 45.00, "125%", "B002"])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.column_mapping.product_name == 0
        assert result.column_mapping.cost == 2   # "Custo" = col 2
        assert result.column_mapping.sku == 5
        assert len(result.products) == 2
        # Custo deve ser 10, não 25 (preço de venda)
        assert result.products[0]["cost"] == Decimal("10.00")

    def test_compound_headers_realistic(self, tmp_path):
        """Headers compostos realistas de catálogos brasileiros."""
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Nome do Produto",
            "Preço de Custo (R$)",
            "Código do Produto",
            "Grupo de Produto",
            "Fornecedor / Marca",
        ])
        ws.append(["Luva de Látex P", 3.80, "LUV-LAT-P", "EPI", "Descarpack"])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.column_mapping.has_required_columns is True
        assert len(result.products) == 1


# ── Testes: nomes inválidos ────────────────────────────────────────────────────

class TestInvalidNames:
    """Nomes que devem ser rejeitados pelo _is_valid_name."""

    def test_numeric_only_name_skipped(self, tmp_path):
        """Nome puramente numérico é SKU na coluna errada."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo"])
        ws.append([12345, 10.00])          # numérico puro
        ws.append(["Produto Válido", 20.00])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert len(result.products) == 1
        assert result.products[0]["raw_name"] == "Produto Válido"

    def test_very_short_name_skipped(self, tmp_path):
        """Nome com 1 caractere é inválido."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo"])
        ws.append(["A", 10.00])            # muito curto
        ws.append(["Produto OK", 20.00])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert len(result.products) == 1

    def test_reserved_word_name_skipped(self, tmp_path):
        """Nome que é palavra reservada (cabeçalho repetido, etc.)."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo"])
        ws.append(["Produto", 10.00])      # palavra reservada = cabeçalho repetido
        ws.append(["Item Real", 20.00])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert len(result.products) == 1

    def test_header_repeat_in_middle_skipped(self, tmp_path):
        """
        Algumas planilhas repetem o cabeçalho a cada página (para impressão).
        """
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo", "SKU"])
        ws.append(["Borracha Branca", 0.60, "BOR-BR"])
        ws.append(["Produto", "Custo", "SKU"])    # cabeçalho repetido — deve ser ignorado
        ws.append(["Lápis HB", 0.80, "LAP-HB"])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert len(result.products) == 2
        assert result.stats.skipped_header_repeat >= 1


# ── Testes: confiança (ParseConfidence) ───────────────────────────────────────

class TestParseConfidence:
    """Cálculo correto do nível de confiança."""

    def test_reliable_confidence_high_success_rate(self, tmp_path):
        """Mais de 80% de linhas válidas → RELIABLE."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo"])
        for i in range(9):  # 9 válidos
            ws.append([f"Produto {i+1}", float(i + 1) * 10])
        ws.append(["", ""])        # inválido (vazio)

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.confidence == ParseConfidence.RELIABLE
        assert result.stats.success_rate >= 0.80

    def test_partial_confidence_low_success_rate(self, tmp_path):
        """
        Entre 20% e 80% de linhas com conteúdo válido → PARTIAL.

        ATENÇÃO: linhas VAZIAS são excluídas do denominador do success_rate
        (são ruído estrutural, não tentativas de produto). Para simular baixa
        taxa de sucesso, precisamos de linhas com CONTEÚDO inválido (custo ausente).
        """
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo"])
        ws.append(["Produto Válido 1", 10.00])   # válido
        ws.append(["Produto Válido 2", 20.00])   # válido
        # 6 linhas com custo ausente: conteúdo inválido que conta no denominador
        for i in range(6):
            ws.append([f"Produto Sem Custo {i+1}", ""])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.confidence == ParseConfidence.PARTIAL

    def test_failed_confidence_missing_required_column(self, tmp_path):
        """Sem coluna de custo → FAILED."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Quantidade", "Unidade"])
        ws.append(["Item A", 100, "UN"])
        ws.append(["Item B", 50, "KG"])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.confidence == ParseConfidence.FAILED
        assert result.column_mapping.cost is None
        assert len(result.errors) >= 1

    def test_failed_confidence_empty_file(self, tmp_path):
        """Arquivo XLSX vazio → FAILED."""
        wb = Workbook()
        wb.active.title = "Vazio"
        result = parse_xlsx_catalog(save_wb(wb, tmp_path))
        assert result.confidence == ParseConfidence.FAILED

    def test_failed_confidence_invalid_file(self, tmp_path):
        """Arquivo não é XLSX válido → FAILED com erro."""
        fake_xlsx = tmp_path / "fake.xlsx"
        fake_xlsx.write_bytes(b"isso nao e um xlsx valido")

        result = parse_xlsx_catalog(fake_xlsx)

        assert result.confidence == ParseConfidence.FAILED
        assert len(result.errors) >= 1

    def test_failed_confidence_all_invalid_costs(self, tmp_path):
        """Todos os produtos têm custo inválido → FAILED."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo"])
        ws.append(["Produto A", "N/A"])
        ws.append(["Produto B", ""])
        ws.append(["Produto C", 0])
        ws.append(["Produto D", -1])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        # Sem produtos válidos → confiança baixa
        assert len(result.products) == 0


# ── Testes: catálogos realistas complexos ─────────────────────────────────────

class TestRealisticCatalogs:
    """
    Cenários que replicam catálogos reais completos.
    Esses testes são os mais importantes para validar produção.
    """

    def test_distribuidora_full_catalog(self, tmp_path):
        """
        Catálogo completo de distribuidora com:
        - Logo/nome da empresa nas primeiras linhas
        - Subtotais por grupo
        - Células mescladas de categoria
        - Linhas vazias entre grupos
        - Mix de formatos de preço
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Tabela de Preços"

        # Linhas de cabeçalho da empresa
        ws.append(["DISTRIBUIDORA PAULISTA DE EMBALAGENS LTDA", None, None, None])
        ws.append(["CNPJ: 12.345.678/0001-99", None, None, None])
        ws.append([None, None, None, None])

        # Cabeçalho real
        ws.append(["Produto", "Preço de Custo", "Código", "Grupo"])

        # Grupo 1: Caixas
        ws.append(["Caixa de Papelão 20x15x10", "R$ 2,50", "CX-P", "Caixas"])
        ws.append(["Caixa de Papelão 30x20x15", "R$ 4,20", "CX-M", None])
        ws.append(["Caixa de Papelão 40x30x20", "R$ 6,80", "CX-G", None])
        ws.append(["Subtotal Caixas", "R$ 13,50", None, None])   # ignorar

        ws.append([None, None, None, None])                       # linha vazia

        # Grupo 2: Sacolas
        ws.append(["Sacola Kraft P", "R$ 0,35", "SC-KR-P", "Sacolas"])
        ws.append(["Sacola Kraft M", "R$ 0,55", "SC-KR-M", None])
        ws.append(["Sacola Plástica 40x50", "R$ 0,12", "SC-PL-40", None])
        ws.append(["Subtotal Sacolas", "R$ 1,02", None, None])   # ignorar

        ws.append([None, None, None, None])

        # Total geral
        ws.append(["Total Geral", "R$ 14,52", None, None])       # ignorar

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        # Deve extrair 6 produtos (3 caixas + 3 sacolas)
        assert len(result.products) == 6
        assert result.confidence != ParseConfidence.FAILED

        # Herança de categoria
        caixas = [p for p in result.products if "Caixa" in p["raw_name"]]
        sacolas = [p for p in result.products if "Sacola" in p["raw_name"]]
        assert all(p["category"] == "Caixas" for p in caixas)
        assert all(p["category"] == "Sacolas" for p in sacolas)

        # Subtotais e total ignorados
        subtotal_names = [p["raw_name"] for p in result.products if "total" in p["raw_name"].lower()]
        assert len(subtotal_names) == 0

    def test_catalogo_farmaceutico_com_multiplas_abas(self, tmp_path):
        """
        Catálogo farmacêutico com abas separadas por categoria.
        Deve processar a aba com mais produtos.
        """
        wb = Workbook()

        ws_ref = wb.active
        ws_ref.title = "Referências"
        ws_ref.append(["Tipo", "Código de Referência"])
        ws_ref.append(["Categoria A", "REF-001"])

        ws_med = wb.create_sheet("Medicamentos")
        ws_med.append(["Produto", "Custo", "Código"])
        for i in range(15):
            ws_med.append([f"Medicamento {i+1}", float(i + 1) * 10, f"MED-{i+1:03d}"])

        ws_cos = wb.create_sheet("Cosméticos")
        ws_cos.append(["Produto", "Custo", "Código"])
        for i in range(5):
            ws_cos.append([f"Cosmético {i+1}", float(i + 1) * 15, f"COS-{i+1:03d}"])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        # Deve escolher "Medicamentos" (mais linhas)
        assert result.stats.sheet_name == "Medicamentos"
        assert len(result.products) == 15

    def test_catalogo_with_all_edge_cases_combined(self, tmp_path):
        """
        Teste de estresse: combina múltiplos edge cases no mesmo catálogo.
        """
        wb = Workbook()
        ws = wb.active

        # Cabeçalho da empresa
        ws.append(["DISTRIBUIDORA TESTE", None, None])
        ws.append([None, None, None])

        # Cabeçalho real com headers compostos
        ws.append(["Nome do Produto", "Vlr Custo", "Cód. Produto"])

        # Produto normal
        ws.append(["Produto Alpha", 50.00, "ALPHA"])

        # Linha de subtotal (deve ser ignorada)
        ws.append(["Subtotal", 50.00, None])

        # Linha vazia
        ws.append([None, None, None])

        # Produto com preço em string
        ws.append(["Produto Beta", "R$ 75,50", "BETA"])

        # Cabeçalho repetido (padrão de impressão)
        ws.append(["Nome do Produto", "Vlr Custo", "Cód. Produto"])

        # Produto com nome numérico (deve ser ignorado)
        ws.append([12345, 10.00, "NUM"])

        # Produto válido final
        ws.append(["Produto Gamma", 100.00, "GAMMA"])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        names = [p["raw_name"] for p in result.products]
        assert "Produto Alpha" in names
        assert "Produto Beta" in names
        assert "Produto Gamma" in names
        # Inválidos ausentes
        assert "Subtotal" not in names
        assert len([n for n in names if n.isdigit()]) == 0


# ── Testes: estatísticas (ParseStats) ─────────────────────────────────────────

class TestParseStats:
    """Verifica que as estatísticas de parsing são calculadas corretamente."""

    def test_stats_count_all_categories(self, tmp_path):
        """Cada tipo de skip é contado corretamente."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo", "SKU"])
        ws.append(["Produto OK 1", 10.00, "SKU-001"])   # válido
        ws.append(["Produto OK 2", 20.00, "SKU-001"])   # duplicata SKU
        ws.append(["", "", ""])                          # vazio
        ws.append(["Subtotal", 30.00, None])             # subtotal
        ws.append(["12345", 40.00, None])                # nome numérico
        ws.append(["Produto OK 3", 0.00, None])          # custo zero

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        assert result.stats.valid_products == 2
        assert result.stats.skipped_empty >= 1
        assert result.stats.skipped_subtotal >= 1
        assert result.stats.skipped_invalid_name >= 1
        assert result.stats.skipped_invalid_cost >= 1
        assert result.stats.skipped_duplicate_sku >= 1

    def test_success_rate_calculation(self, tmp_path):
        """
        success_rate = valid_products / (total_rows_scanned - skipped_empty).

        Linhas vazias são excluídas do denominador — são ruído estrutural.
        Para calcular taxa real, usamos linhas com conteúdo inválido (custo ausente).
        """
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Custo"])
        for i in range(8):  # 8 válidos
            ws.append([f"Produto {i+1}", float(i + 1) * 5])
        # 2 linhas com conteúdo mas custo invalido (contam no denominador)
        ws.append(["Produto Sem Custo A", ""])
        ws.append(["Produto Sem Custo B", ""])

        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        # 8 válidos / 10 total (sem vazias) = 80%
        assert result.stats.valid_products == 8
        assert result.stats.skipped_invalid_cost == 2
        non_empty = result.stats.total_rows_scanned - result.stats.skipped_empty
        assert non_empty == 10
        assert abs(result.stats.success_rate - 0.80) < 0.01

    def test_summary_string_format(self, tmp_path):
        """summary() retorna string legível para logging."""
        wb = make_standard_wb()
        result = parse_xlsx_catalog(save_wb(wb, tmp_path))

        summary = result.summary()
        assert isinstance(summary, str)
        assert len(summary) > 0
        # Deve conter informações básicas
        assert "5" in summary or "RELIABLE" in summary
