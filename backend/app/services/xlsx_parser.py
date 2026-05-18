"""
XLSX Parser — Extração robusta de produtos de catálogos reais.

Projetado para lidar com a realidade dos catálogos brasileiros:
- Cabeçalhos em qualquer linha (não necessariamente a primeira)
- Nomes de colunas inconsistentes entre fornecedores
- Linhas de subtotal misturadas com produtos
- Células mescladas para categorias
- Preços em qualquer formato (R$ 1.234,56, 1234.56, "1.234", etc.)
- Planilhas com múltiplas abas
- Linhas vazias no meio dos dados
- SKUs com zeros à esquerda (perdidos se célula é numérica)

Uso:
    parser = XLSXParser(file_path=Path("catalogo.xlsx"))
    result = parser.parse()
    print(result.summary())
    for product in result.products:
        print(product["raw_name"], product["cost"])
"""

import logging
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.services.column_detector import detect_columns, detect_header_row, normalize_text
from app.services.parse_result import (
    ColumnMappingResult,
    ParseConfidence,
    ParseResult,
    ParseStats,
)
from app.services.price_normalizer import is_valid_cost, normalize_price

logger = logging.getLogger(__name__)

# ── Configurações ─────────────────────────────────────────────────────────────

# Máximo de linhas vazias consecutivas antes de parar de ler
MAX_CONSECUTIVE_EMPTY = 10

# Linhas iniciais para escanear buscando cabeçalho
MAX_HEADER_SCAN_ROWS = 12

# Keywords que indicam linhas de total/subtotal
SUBTOTAL_KEYWORDS = frozenset([
    "total", "subtotal", "soma", "sum", "grand total", "total geral",
    "total parcial", "totalização", "totalizacao", "total de itens",
])

# Keywords que indicam que o nome é inválido (erro de parsing, não produto)
INVALID_NAME_KEYWORDS = frozenset([
    "total", "subtotal", "soma", "item", "produto", "nome",  # cabeçalho repetido
    "descricao", "descrição", "n/a", "none", "null", "-", "#",
])

# Tamanho mínimo/máximo de nome de produto
MIN_NAME_LENGTH = 2
MAX_NAME_LENGTH = 300

# Custo máximo razoável para um produto de catálogo (em BRL)
MAX_REASONABLE_COST = 999_999


class XLSXParser:
    """
    Parser de catálogos XLSX com detecção inteligente de estrutura.

    Cada instância processa um único arquivo.
    Imutável após parse() — cria novo XLSXParser para novo arquivo.
    """

    def __init__(self, file_path: Path) -> None:
        self.file_path = file_path
        self._result = ParseResult()
        self._stats = ParseStats()
        self._seen_skus: set[str] = set()  # Detectar duplicatas dentro do catálogo

    def parse(self) -> ParseResult:
        """
        Executa o parsing completo do arquivo XLSX.

        Returns:
            ParseResult com produtos, confiança e estatísticas
        """
        logger.info("XLSXParser: iniciando | arquivo=%s", self.file_path.name)

        try:
            wb = openpyxl.load_workbook(
                str(self.file_path),
                read_only=True,
                data_only=True,  # Retorna valores calculados, não fórmulas
            )
        except Exception as exc:
            self._result.add_error(f"Não foi possível abrir o arquivo: {exc}")
            self._result.confidence = ParseConfidence.FAILED
            logger.error("XLSXParser: falha ao abrir arquivo: %s", exc)
            return self._result

        try:
            # Escolher a melhor aba para processar
            ws, sheet_name = self._select_best_sheet(wb)
            self._stats.sheet_name = sheet_name
            self._stats.total_sheets = len(wb.sheetnames)

            if ws is None:
                self._result.add_error("Nenhuma aba com dados encontrada")
                self._result.confidence = ParseConfidence.FAILED
                return self._result

            logger.info("XLSXParser: processando aba '%s'", sheet_name)

            # Carregar todas as linhas em memória (necessário para detecção de cabeçalho)
            # Para arquivos muito grandes (>50k linhas), isso pode ser problemático
            # Fase 2: streaming com janela deslizante
            all_rows = self._load_rows(ws)

            if not all_rows:
                self._result.add_error("Planilha vazia")
                self._result.confidence = ParseConfidence.FAILED
                return self._result

            # Detectar linha de cabeçalho
            header_idx, header_confidence = detect_header_row(all_rows, MAX_HEADER_SCAN_ROWS)
            self._stats.header_row_index = header_idx

            if header_confidence < 0.3:
                self._result.add_warning(
                    f"Linha de cabeçalho detectada com baixa confiança ({header_confidence:.0%}). "
                    f"Usando linha {header_idx + 1}."
                )
            logger.info(
                "XLSXParser: cabeçalho na linha %d (confiança %.0f%%)",
                header_idx + 1, header_confidence * 100
            )

            # Mapear colunas
            header_row = all_rows[header_idx]
            col_mapping = detect_columns(header_row)
            self._result.column_mapping = col_mapping
            self._stats.missing_optional_cols = col_mapping.missing_optional

            self._log_column_mapping(col_mapping)

            if not col_mapping.has_required_columns:
                missing = []
                if col_mapping.product_name is None:
                    missing.append("nome do produto")
                if col_mapping.cost is None:
                    missing.append("custo/preço")

                self._result.add_error(
                    f"Colunas obrigatórias não encontradas: {', '.join(missing)}. "
                    f"Cabeçalhos detectados: {[h for h in col_mapping.original_headers if h]}"
                )
                self._result.confidence = ParseConfidence.FAILED
                self._result.stats = self._stats
                return self._result

            # Processar linhas de dados (abaixo do cabeçalho)
            data_rows = all_rows[header_idx + 1:]
            products = self._process_data_rows(data_rows, col_mapping)

            self._result.products = products
            self._result.stats = self._stats
            self._result.confidence = self._calculate_confidence(col_mapping)

            logger.info("XLSXParser: %s", self._result.summary())

        finally:
            wb.close()

        return self._result

    # ── Seleção de aba ────────────────────────────────────────────────────────

    def _select_best_sheet(self, wb) -> tuple[Worksheet | None, str]:
        """
        Escolhe a aba com mais dados.

        Estratégia:
        1. Se há aba ativa com dados, usar ela
        2. Caso contrário, escolher a aba com mais linhas não-vazias
        3. Ignorar abas com nomes que sugerem auxiliar ("config", "lookup", "ref")
        """
        SKIP_SHEET_NAMES = frozenset([
            "config", "configuracao", "configuração", "setup",
            "ref", "reference", "lookup", "auxiliar", "aux",
            "legenda", "legend", "instrucoes", "instruções",
        ])

        candidates: list[tuple[str, int]] = []  # (sheet_name, row_count)

        for sheet_name in wb.sheetnames:
            norm_name = normalize_text(sheet_name)

            # Pular abas auxiliares
            if any(skip in norm_name for skip in SKIP_SHEET_NAMES):
                logger.debug("XLSXParser: ignorando aba '%s' (auxiliar)", sheet_name)
                continue

            ws = wb[sheet_name]
            # Contar linhas não-vazias (até 5 para ser rápido)
            non_empty = sum(1 for row in ws.iter_rows(max_row=1000, values_only=True) if any(row))
            candidates.append((sheet_name, non_empty))

        if not candidates:
            # Fallback: usar primeira aba
            first = wb.sheetnames[0] if wb.sheetnames else None
            if first:
                return wb[first], first
            return None, ""

        # Ordenar por contagem de linhas desc
        candidates.sort(key=lambda x: x[1], reverse=True)
        best_name = candidates[0][0]

        if len(candidates) > 1:
            self._result.add_warning(
                f"Múltiplas abas encontradas: {[c[0] for c in candidates]}. "
                f"Processando '{best_name}' (maior quantidade de dados)."
            )

        return wb[best_name], best_name

    # ── Carregamento de linhas ────────────────────────────────────────────────

    def _load_rows(self, ws: Worksheet) -> list[list[Any]]:
        """
        Carrega todas as linhas da planilha como lista de listas.

        Trata células mescladas: openpyxl retorna None para células não-master
        de merged ranges em read_only mode, então simplesmente mantemos None.
        """
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        return rows

    # ── Processamento de linhas de dados ─────────────────────────────────────

    def _process_data_rows(
        self, data_rows: list[list], col_mapping: ColumnMappingResult
    ) -> list[dict]:
        """
        Processa cada linha de dados e extrai produtos válidos.

        Mantém contexto de categoria para herança de células mescladas.
        """
        products: list[dict] = []
        consecutive_empty = 0
        last_valid_category: str | None = None  # Herança de categoria de células mescladas

        for row_idx, row in enumerate(data_rows):
            self._stats.total_rows_scanned += 1

            # Verificar linha vazia
            if self._is_empty_row(row):
                self._stats.skipped_empty += 1
                consecutive_empty += 1
                if consecutive_empty >= MAX_CONSECUTIVE_EMPTY:
                    logger.debug(
                        "XLSXParser: %d linhas vazias consecutivas — parando na linha %d",
                        MAX_CONSECUTIVE_EMPTY, row_idx + 1
                    )
                    self._stats.total_rows_scanned -= MAX_CONSECUTIVE_EMPTY
                    break
                continue

            consecutive_empty = 0  # Reset contador

            # Verificar se é linha de subtotal
            if self._is_subtotal_row(row, col_mapping):
                self._stats.skipped_subtotal += 1
                logger.debug("XLSXParser: linha %d ignorada (subtotal)", row_idx + 1)
                continue

            # Verificar se é repetição do cabeçalho
            if self._is_header_repeat(row, col_mapping):
                self._stats.skipped_header_repeat += 1
                continue

            # Extrair produto
            product, skip_reason = self._extract_product(row, col_mapping, last_valid_category)

            if product is None:
                self._record_skip(skip_reason, row_idx + 1, row, col_mapping)
                continue

            # Herdar categoria para próximas linhas (padrão de células mescladas)
            if product.get("category"):
                last_valid_category = product["category"]

            # Verificar duplicata de SKU
            if product.get("sku"):
                sku_norm = str(product["sku"]).strip().upper()
                if sku_norm in self._seen_skus:
                    self._stats.skipped_duplicate_sku += 1
                    logger.debug(
                        "XLSXParser: SKU duplicado '%s' na linha %d — ignorado",
                        sku_norm, row_idx + 1
                    )
                    # Não skip completo — pode ser produto diferente com mesmo SKU
                    # Manter mas logar
                    self._result.add_warning(
                        f"SKU duplicado: '{sku_norm}' aparece mais de uma vez no catálogo"
                    )
                self._seen_skus.add(sku_norm)

            products.append(product)
            self._stats.valid_products += 1

        logger.info(
            "XLSXParser: %d produtos válidos de %d linhas scaneadas",
            len(products), self._stats.total_rows_scanned
        )

        return products

    # ── Extração de produto de uma linha ─────────────────────────────────────

    def _extract_product(
        self,
        row: list,
        col_mapping: ColumnMappingResult,
        inherited_category: str | None,
    ) -> tuple[dict | None, str]:
        """
        Extrai um produto de uma linha.

        Returns:
            (product_dict, skip_reason) — skip_reason="" se produto é válido
        """
        # Nome do produto (obrigatório)
        raw_name_val = self._get_cell(row, col_mapping.product_name)
        raw_name = self._clean_text(raw_name_val)

        if not self._is_valid_name(raw_name):
            return None, "invalid_name"

        # Custo (obrigatório)
        cost_val = self._get_cell(row, col_mapping.cost)
        cost = normalize_price(cost_val)

        if cost is None:
            self._stats.cost_parse_errors += 1
            return None, "invalid_cost"

        if not is_valid_cost(cost):
            return None, "invalid_cost"

        # Campos opcionais
        sku = self._clean_sku(self._get_cell(row, col_mapping.sku))
        category = self._clean_text(self._get_cell(row, col_mapping.category)) or inherited_category
        supplier = self._clean_text(self._get_cell(row, col_mapping.supplier))

        return {
            "raw_name": raw_name,
            "cost": cost,
            "sku": sku,
            "category": category,
            "supplier": supplier,
            "currency": "BRL",
        }, ""

    # ── Detecção de linhas especiais ─────────────────────────────────────────

    def _is_empty_row(self, row: list) -> bool:
        """Retorna True se a linha não tem nenhum valor não-nulo/não-vazio."""
        return not any(
            cell is not None and str(cell).strip()
            for cell in row
        )

    def _is_subtotal_row(self, row: list, col_mapping: ColumnMappingResult) -> bool:
        """
        Detecta linhas de total/subtotal.

        Heurísticas:
        1. Primeira célula não-vazia contém keyword de total
        2. Célula de nome contém keyword de total
        3. Nome contém apenas palavras-chave de total
        """
        # Verificar primeira célula não-vazia
        for cell in row:
            if cell is not None and str(cell).strip():
                cell_norm = normalize_text(str(cell))
                for kw in SUBTOTAL_KEYWORDS:
                    if kw in cell_norm:
                        return True
                break  # Só verificar a primeira célula não-vazia

        # Verificar célula específica de nome
        if col_mapping.product_name is not None:
            name_val = self._get_cell(row, col_mapping.product_name)
            if name_val:
                name_norm = normalize_text(str(name_val))
                for kw in SUBTOTAL_KEYWORDS:
                    if name_norm.startswith(kw) or name_norm == kw:
                        return True

        return False

    def _is_header_repeat(self, row: list, col_mapping: ColumnMappingResult) -> bool:
        """
        Detecta se a linha é uma repetição do cabeçalho.

        Isso ocorre em planilhas com cabeçalho repetido a cada N linhas (impressão).
        """
        if col_mapping.product_name is None:
            return False

        name_val = self._get_cell(row, col_mapping.product_name)
        if not name_val:
            return False

        name_norm = normalize_text(str(name_val))
        return name_norm in INVALID_NAME_KEYWORDS

    # ── Utilitários de célula ─────────────────────────────────────────────────

    def _get_cell(self, row: list, col_index: int | None) -> Any:
        """Acessa célula por índice com segurança (retorna None se fora do range)."""
        if col_index is None or col_index >= len(row):
            return None
        return row[col_index]

    def _clean_text(self, value: Any) -> str | None:
        """Limpa valor de célula de texto: strip, remove quebras de linha, normaliza espaços."""
        if value is None:
            return None
        text = str(value).strip()
        # Remove quebras de linha internas (comum em células editadas manualmente)
        text = re.sub(r"[\r\n\t]+", " ", text)
        # Normaliza espaços múltiplos
        text = " ".join(text.split())
        if not text or text.lower() in ("nan", "none", "null", "-", "n/a", "#n/d", "#ref!"):
            return None
        return text

    def _clean_sku(self, value: Any) -> str | None:
        """
        Limpa SKU mantendo zeros à esquerda.

        openpyxl retorna inteiros para células numéricas — zeros à esquerda são perdidos.
        Ex: SKU "00123" → célula numérica 123 → recuperamos como "123" (sem os zeros)
        Isso é uma limitação do Excel. Workaround: formatar célula como texto no Excel.
        """
        if value is None:
            return None

        # Se é número, converter para string sem decimais
        if isinstance(value, (int, float)):
            if value == int(value):
                return str(int(value))
            return str(value)

        text = self._clean_text(value)
        if not text:
            return None

        # Remover caracteres problemáticos mas manter zeros à esquerda
        text = text.strip()
        return text if text else None

    def _is_valid_name(self, name: str | None) -> bool:
        """
        Valida se o nome do produto é aceitável.

        Rejeita:
        - None ou vazio
        - Muito curto (provavelmente código ou erro)
        - Muito longo (provavelmente descrição errada)
        - Só números (provavelmente SKU na coluna errada)
        - Palavras reservadas (total, subtotal, etc.)
        """
        if not name:
            return False

        name_stripped = name.strip()

        if len(name_stripped) < MIN_NAME_LENGTH:
            return False

        if len(name_stripped) > MAX_NAME_LENGTH:
            return False

        # Nome é só números (SKU ou código)
        if re.match(r"^[\d\s\.\-/]+$", name_stripped):
            return False

        # Palavra reservada
        name_norm = normalize_text(name_stripped)
        if name_norm in INVALID_NAME_KEYWORDS:
            return False

        return True

    # ── Confiança e logging ───────────────────────────────────────────────────

    def _calculate_confidence(self, col_mapping: ColumnMappingResult) -> ParseConfidence:
        """
        Calcula nível de confiança baseado em:
        - Colunas obrigatórias encontradas
        - Taxa de sucesso de linhas
        - Colunas opcionais encontradas
        """
        if not col_mapping.has_required_columns:
            return ParseConfidence.FAILED

        success_rate = self._stats.success_rate

        if success_rate >= 0.80:
            return ParseConfidence.RELIABLE
        elif success_rate >= 0.20:
            return ParseConfidence.PARTIAL
        else:
            if self._stats.valid_products == 0:
                return ParseConfidence.FAILED
            return ParseConfidence.PARTIAL

    def _log_column_mapping(self, col_mapping: ColumnMappingResult) -> None:
        """Loga o mapeamento de colunas de forma legível."""
        found = []
        missing = []

        mapping = {
            "nome": col_mapping.product_name,
            "custo": col_mapping.cost,
            "sku": col_mapping.sku,
            "categoria": col_mapping.category,
            "fornecedor": col_mapping.supplier,
        }

        for field_label, col_idx in mapping.items():
            if col_idx is not None:
                original = col_mapping.original_headers[col_idx] if col_idx < len(col_mapping.original_headers) else "?"
                score = col_mapping.scores.get(
                    {"nome": "product_name", "custo": "cost", "sku": "sku",
                     "categoria": "category", "fornecedor": "supplier"}[field_label],
                    0.0
                )
                found.append(f"{field_label}=col[{col_idx}]:'{original}'({score:.0%})")
            else:
                missing.append(field_label)

        logger.info("XLSXParser: colunas mapeadas → %s", " | ".join(found))
        if missing:
            logger.warning("XLSXParser: colunas não encontradas → %s", ", ".join(missing))

        # Adicionar warning para colunas opcionais ausentes
        for col in col_mapping.missing_optional:
            self._result.add_warning(f"Coluna opcional '{col}' não encontrada no catálogo")

        # Adicionar info sobre colunas não reconhecidas
        # (calculada no detect_columns mas precisamos recuperar)

    def _record_skip(
        self,
        reason: str,
        row_num: int,
        row: list,
        col_mapping: ColumnMappingResult,
    ) -> None:
        """Registra motivo de skip e atualiza estatísticas."""
        if reason == "invalid_name":
            self._stats.skipped_invalid_name += 1
            name_val = self._get_cell(row, col_mapping.product_name)
            logger.debug(
                "XLSXParser: linha %d ignorada (nome inválido): '%s'",
                row_num, name_val
            )
        elif reason == "invalid_cost":
            self._stats.skipped_invalid_cost += 1
            cost_val = self._get_cell(row, col_mapping.cost)
            logger.debug(
                "XLSXParser: linha %d ignorada (custo inválido): '%s'",
                row_num, cost_val
            )


# ── Função de conveniência ────────────────────────────────────────────────────

def parse_xlsx_catalog(file_path: Path) -> ParseResult:
    """
    Função de conveniência que instancia e executa o parser.

    Args:
        file_path: Caminho para o arquivo .xlsx

    Returns:
        ParseResult com produtos e diagnóstico
    """
    parser = XLSXParser(file_path)
    return parser.parse()
